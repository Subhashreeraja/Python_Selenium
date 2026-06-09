import openpyxl

def get_data(path,sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    row = sheet.max_row
    column = sheet.max_column

    final_list = []
    for r in range(2, row + 1):
        row_list = []
        for c in range(1, column + 1):
            row_list.append(sheet.cell(r,c).value)
        final_list.append(row_list)
    
    return final_list      